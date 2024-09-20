from sqlalchemy import create_engine
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.keys import Keys
from time import sleep
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import schedule
import json

def connect_db():
    with open('keys.json', 'r') as file:
        data = json.load(file)

    host, database = data['host'], data['database']

    connection_string = f'mssql+pyodbc://{host}/{database}?driver=ODBC+Driver+17+for+SQL+Server'

    try:
        engine = create_engine(connection_string)
        print("Conexion exitosa")
    except Exception as e:
        print("Error al conectar")
    
    return engine

def get_data(engine,written_query):
    query = f"{written_query}"
    df = pd.read_sql(query, engine)
    engine.dispose() 
    return df

    
def clean_data(df):
    df = df.sort_values(by=['nombre_sede','fechahora'],ascending=True).reset_index(drop=True)
    # df.rename(columns={'fechahora_rounded':'date_time','valor_mean':'valor'},inplace=True)
    df['fechahora'] = pd.to_datetime(df['fechahora'])
    return df

def connect_wsp():
    
    with open('keys.json', 'r') as file:
        data = json.load(file)
        
    opts = Options()
    driver = webdriver.Edge(
        service = Service(EdgeChromiumDriverManager().install()),
        options = opts
    )
    driver.get('https://web.whatsapp.com/')
    sleep(60)
    titulo_chat =  data['titulo_chat']  ## Cambiar

    try:
        # Intenta encontrar el chat por su título
        chat = driver.find_element(By.XPATH, f'//span[@title="{titulo_chat}"]')
        chat.click()
    except:
        # Si no se encuentra el chat, imprime un mensaje y maneja la excepción como desees
        print("No se encontró el chat con el título proporcionado.")
    return driver


def show_txt(tipo,value,campo_texto, bool_sub=0):
    # Nombre de la sede
    if(bool_sub==1):
        campo_texto.send_keys(">")
        campo_texto.send_keys(Keys.SPACE)

    header = f"{tipo} {value}"
    campo_texto.send_keys(header)
    campo_texto.send_keys( Keys.SHIFT + Keys.ENTER)

def message_T(df,data,warning,alert):
    for id,row in df.iterrows():
        if((row['valor'] < 22) & (row['valor']> 20.5)):
            warning = warning +1 
        elif(row['valor'] >= 22):
            alert = alert + 1

    for id,row in df.iterrows():

        if((row['valor'] < 22) & (row['valor']> 20.5)):
            data.append({"Tipo de Alerta": "Warning","Variable":"Temperatura", "Sede": row['nombre_sede'], "Zona": row['nombre_zona'], "Valor": row['valor'], "Hora":row['fechahora'].strftime('%H:%M:%S')})

        elif(row['valor'] >= 22):
            data.append({"Tipo de Alerta": "Alert","Variable":"Temperatura", "Sede": row['nombre_sede'], "Zona": row['nombre_zona'], "Valor": row['valor'], "Hora":row['fechahora'].strftime('%H:%M:%S')})
    
    return data, warning,alert


def message_H(df,driver,data,warning,alert):
    campo_texto = driver.find_element(By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')

    for id,row in df.iterrows():
        if((row['valor'] < 70) & (row['valor']>= 65)):
            warning = warning +1 
        elif(row['valor'] >= 70):
            alert = alert + 1

    total = alert + warning

    if(total ==0):
        campo_texto.send_keys(":perfect")
        campo_texto.send_keys(Keys.ENTER)
        campo_texto.send_keys(Keys.SPACE)
        campo_texto.send_keys('No hay alertas')
        boton_enviar = driver.find_element(By.XPATH, '//span[@data-icon="send"]')
        boton_enviar.click()

    else:
        campo_texto.send_keys(":warning")
        campo_texto.send_keys(Keys.ENTER)
        campo_texto.send_keys(Keys.SPACE)
        sleep(1)
        show_txt('Warning:',warning,campo_texto)

        campo_texto.send_keys(":alert")
        campo_texto.send_keys(Keys.ENTER)
        campo_texto.send_keys(Keys.SPACE)
        sleep(1)
        show_txt('Alert:',alert,campo_texto)
        
        campo_texto.send_keys(Keys.SHIFT + Keys.ENTER)
        boton_enviar = driver.find_element(By.XPATH, '//span[@data-icon="send"]')
        boton_enviar.click()

        for id,row in df.iterrows():

            if((row['valor'] < 70) & (row['valor']>= 65)):
                data.append({"Tipo de Alerta": "Warning","Variable":"Humedad", "Sede": row['nombre_sede'], "Zona": row['nombre_zona'], "Valor": row['valor'], "Hora":row['fechahora'].strftime('%H:%M:%S')})

            elif(row['valor'] >= 70):
                data.append({"Tipo de Alerta": "Alert","Variable":"Humedad", "Sede": row['nombre_sede'], "Zona": row['nombre_zona'], "Valor": row['valor'],"Hora":row['fechahora'].strftime('%H:%M:%S')})
    
    return data, total

def create_excel(data):

    data = pd.DataFrame(data,index=None)
    nombre_archivo = 'reporte.xlsx'
    book = Workbook()

    # Crear hoja de trabajo
    sheet = book.active

    # Definir los colores para el formato condicional
    fill_rojo = PatternFill(start_color="CC5555", end_color="CC5555", fill_type="solid")
    fill_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Escribir los encabezados
    for col_idx, col_name in enumerate(data.columns, start=1):
        sheet.cell(row=1, column=col_idx).value = col_name

    # Escribir los datos
    for row_idx, row_data in enumerate(data.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value

    # Aplicar formato condicional a la columna 'Tipo de Alerta'
    for id, value in enumerate(data['Tipo de Alerta'], start=2):  # Iniciar desde la segunda fila
        if value == "Warning":
            sheet.cell(row=id, column=1).fill = fill_amarillo  # Columna 'nombre_alternativo'
            sheet.cell(row=id, column=5).fill = fill_amarillo  # Columna 'nombre_alternativo'
        else:
            sheet.cell(row=id, column=1).fill = fill_rojo  # Columna 'nombre_alternativo'
            sheet.cell(row=id, column=5).fill = fill_rojo  # Columna 'nombre_alternativo'

    book.save(nombre_archivo)



def send_excel(driver,data):
    create_excel(data)
    file_path= r"C:\Users\OPERACIONES\Documents\Archivos VS Code\Alertas_por_wsp\reporte.xlsx"
    # Find the paperclip icon (attachment button)
    attachment_button = driver.find_element(By.XPATH, "//div[@title='Adjuntar']")
    attachment_button.click()
    sleep(3)
    document_upload_option = driver.find_element(By.XPATH, "//input[@type='file']")  # Adjust XPath if needed
    document_upload_option.send_keys(file_path)
    sleep(3)

    campo_texto = driver.find_element(By.XPATH, '//div[@contenteditable="true"]')
    sleep(5)
    campo_texto.send_keys(Keys.ENTER)
    sleep(5)


def at_hour(function):

    schedule.every().day.at("05:00:00").do(function)
    schedule.every().day.at("05:30:00").do(function)
    schedule.every().day.at("06:00:00").do(function)
    schedule.every().day.at("06:30:00").do(function)
    schedule.every().day.at("07:00:00").do(function)
    schedule.every().day.at("07:30:00").do(function)
    schedule.every().day.at("08:00:00").do(function)
    schedule.every().day.at("08:30:00").do(function)
    schedule.every().day.at("09:00:00").do(function)
    schedule.every().day.at("09:30:00").do(function)
    schedule.every().day.at("10:00:00").do(function)
    schedule.every().day.at("10:30:00").do(function)
    schedule.every().day.at("11:00:00").do(function)
    schedule.every().day.at("11:30:00").do(function)
    schedule.every().day.at("12:00:00").do(function)
    schedule.every().day.at("12:30:00").do(function)
    schedule.every().day.at("13:00:00").do(function)
    schedule.every().day.at("13:30:00").do(function)
    schedule.every().day.at("14:00:00").do(function)
    schedule.every().day.at("14:30:00").do(function)
    schedule.every().day.at("15:00:00").do(function)
    schedule.every().day.at("15:30:00").do(function)
    schedule.every().day.at("16:00:00").do(function)
    schedule.every().day.at("16:30:00").do(function)
    schedule.every().day.at("17:00:00").do(function)
    schedule.every().day.at("17:30:00").do(function)
    schedule.every().day.at("18:00:00").do(function)
    schedule.every().day.at("18:30:00").do(function)
    schedule.every().day.at("19:00:00").do(function)
    schedule.every().day.at("19:30:00").do(function)
    schedule.every().day.at("20:00:00").do(function)
    schedule.every().day.at("20:30:00").do(function)
    schedule.every().day.at("21:00:00").do(function)
    schedule.every().day.at("21:30:00").do(function)
    schedule.every().day.at("22:00:00").do(function)
    schedule.every().day.at("22:30:00").do(function)
    schedule.every().day.at("23:00:00").do(function)

    while True:
        schedule.run_pending()
        sleep(1)